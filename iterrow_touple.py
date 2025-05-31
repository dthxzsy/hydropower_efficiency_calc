import pandas as pd
import openpyxl
from openpyxl.styles import Font, Side, Alignment, Border

# Create a DataFrame
df = pd.DataFrame({
    "Name": ["Bob", "Jack", "Hely"],
    "Age": [20, 23, 17],
    "Score": [87, 56, 92]
})

# Create a new workbook and select the active worksheet
wb = openpyxl.Workbook()
ws = wb.active

# Define styles
font = Font(name="Calibri", size=11, bold=True)
alignment = Alignment(horizontal="center", vertical="center")
thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin")
)

for col_idx, col_name in enumerate(df.columns, start=1):
    cell = ws.cell(row=1, column=col_idx, value=col_name)
    cell.font = font
    cell.alignment = alignment
    cell.border = thin_border

for row_idx, row in enumerate(df.itertuples(index=False), start=2):
    for col_idx, value in enumerate(row, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.font = font
        cell.alignment = alignment
        cell.border = thin_border
# Save the workbook
wb.save(r"C:\Users\Administrator\Desktop\演示\styled_output.xlsx")
