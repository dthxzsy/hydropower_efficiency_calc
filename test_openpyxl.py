import openpyxl
from openpyxl.styles import Font, Alignment
master_data = openpyxl.load_workbook(r'C:\Users\Administrator\Desktop\project_demo\master_sheet.xlsx')
daily_data = openpyxl.load_workbook(r'C:\Users\Administrator\Desktop\project_demo\daily_sheet.xlsx')

master_sheet = master_data['data']
daily_sheet = daily_data['Sheet1']

is_data = True
master_count = 1
while is_data:
    master_count += 1
    data = master_sheet.cell(row=master_count, column=1).value
    if data is None:
        is_data = False

is_data = True
daily_count = 1
while is_data:
    daily_count += 1
    data = daily_sheet.cell(row=daily_count, column=1).value
    if data is None:
        is_data = False
# print(master_count)
# print(daily_count)

for i in range(1,daily_count):
    ID = daily_sheet.cell(row=i, column=1).value
    row_num = daily_sheet.cell(row=i, column=2).row
    for j in range(1,master_count):
        if master_sheet.cell(row=j, column=1).value == ID:
            Todays_Purchase = daily_sheet.cell(row=row_num, column=2).value
            Todays_Reward = daily_sheet.cell(row=row_num, column=3).value
            

            Total_Purchases = master_sheet.cell(row=j, column=6).value
            Life_Time_Reward_Balance = master_sheet.cell(row=j, column=7).value


            new_Total_Purchases = Total_Purchases + Todays_Purchase
            new_Life_Time_Reward_Balance = Life_Time_Reward_Balance + Todays_Reward

            master_sheet.cell(row=j, column=6).value = new_Total_Purchases
            master_sheet.cell(row=j, column=7).value = new_Life_Time_Reward_Balance

master_data.save(r'C:\Users\Administrator\Desktop\project_demo\updated_master_sheet.xlsx')


report_data = openpyxl.Workbook()
ws=report_data.active
head_style = Font(name = 'Times New Roman', size = 12, bold=True)
header_item = []
colun_count = 1

while True:
    cell_value = master_sheet.cell(row=1, column=colun_count).value
    if cell_value is None:
        break
    header_item.append(cell_value)
    colun_count += 1
    
for index, data in enumerate(header_item):
    colun_count = index + 1
    cell = ws.cell(row=1, column=colun_count)  # Using ws here instead of master_sheet
    cell.font = head_style
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.value = data

print(todays)










